import openpyxl as xl
import os
import sqlite3
import json
from pathlib import Path
import pandas as pd
import numpy  as np
import math

#parse json
def generate_create_command(db_dct):
    """
    Generates a list of SQL CREATE TABLE commands based on the provided database dictionary.

    Args:
        db_dct (dict): A dictionary representing the database schema, where each key is a table name and each value is another dictionary containing column names and their respective data types.

    Returns:
        list: A list of SQL CREATE TABLE commands as strings.
    """
    
    base_string = []
    for table_name in db_dct:
        base_create_table = f"CREATE TABLE {table_name} ("

        for column_name in db_dct[table_name]:
            base_create_table += f"{column_name} {db_dct[table_name][column_name]}, "

        base_create_table = base_create_table[:-2] + ");"
        base_string.append(base_create_table)
    
    return base_string



def generate_database(path, out_db):
    """
	Generates a SQLite database based on a provided JSON schema.

	Args:
		path (str): The path to the JSON schema file.
		out_db (str): The path to the output SQLite database file.

	Returns:
		None
	"""

    #Check if database exists
    if os.path.exists(out_db):
        print("DB already exists, deleting file...")
        os.remove(out_db)
        print('Deleted...')

    print('loading json...')
    #Loads Data base
    with open(path, 'r') as f:
        db_dct = json.load(f)

    #Creates the database
    print('Creating DB...')
    conn = sqlite3.connect(out_db)
    cursor = conn.cursor()

    #Creates the table commands and exectutes
    base_create_string = generate_create_command(db_dct)
    for create_command in base_create_string:
        cursor.execute(create_command)
    conn.commit()
    conn.close()

    print(f"Created DB: {out_db}")

class FileManager:
    """File Manager for the program"""
    
    def __init__(self):        
        self.storage = {'display_name':[],
                        'path':[],
                        'type':[]}
        
    def quick_append(self, file_path, type):
        self.storage['display_name'].append(os.path.basename(file_path))
        self.storage['path'].append(Path(file_path))
        self.storage['type'].append(type)
    
    def quick_remove(self, idx):
        self.storage['display_name'].pop(idx)
        self.storage['path'].pop(idx)
        self.storage['type'].pop(idx)

    def get_storage(self):
        return self.storage
    
    def add_pcef(self, file_path):
        self.storage['pcef'] = file_path
    
    def add_capacity(self, file_path):
        self.storage['capacity'] = file_path
    def reset(self):
        self.storage = {'display_name':[],
                        'path':[],
                        'type':[]}

def file_check(file:Path|str, type:str):
        """Returns True if file is valid, False if not
        
        Check Criteria:
        
        
            MB: 
                Cell['A1'] == 'Project ID'
                Cell['A2'] == 'Time Start'
                Cell['B2'] == 'Time End'
                
            3LI:    
                Cell['A2'] == 'TIME'
                Cell['B2'] == 'Start Hour'
                Cell['D2'] == 'End Hour'

            4LI:    
                Cell['A1'] == 'TIME'
                Cell['B1'] == 'Start Hour'
                Cell['D1'] == 'End Hour'
                
            SS: CurrentlyWorking
            """
        
        file = xl.load_workbook(file)
        file = file.active

        if type == 'MB':
            p_id = file['A1'].value
            t_start = file['A2'].value
            t_end = file['B2'].value

            if (p_id == 'Project ID') and (t_start == 'Time Start') and (t_end == 'Time End'):
                return True
            else:
                return False
        elif type == '3LI':
            t_start = file['B2'].value
            t_end = file['D2'].value
            time_text = file['A2'].value

            if (t_start == 'Start Hour') and (t_end == 'End Hour') and (time_text == 'TIME'):
                return True
            else:
                return False
        
        elif type == '4LI':
            t_start = file['B1'].value
            t_end = file['D1'].value
            time_text = file['A1'].value

            if (t_start == 'Start Hour') and (t_end == 'End Hour') and (time_text == 'TIME'):
                return True
            else:
                return False
        
        else:
            raise ValueError("Invalid file type")


def fourleg_intersection_to_db(wb_path, database_path):
    """
    This function reads data from 4-Leg intersection Excel file and inserts it into a SQLite database.

    Parameters:
        wb_path (str): The path to the Excel file.
        database_path (str): The path to the SQLite database.

    Returns:
        None
    """
    wb = xl.load_workbook(wb_path)
    ws = wb.active


    #Get the start Hour
    start_hour  = ws['C1']
    end_hour    = ws['E1']
    project_id  = ws['L1'].value
    date        = ws['H1']
    approach    = [i.value for i in ws['B2:K2'][0] if i.value != None]
    egress      = [i.value for i in ws['B3:M3'][0]]
    vehicles    = [[i[0].value for i in ws['A4:A11']]]
    data        = ws['B4:M11']

    #Transforms the openpyxl cell to actual data
    data_arr = []
    for row in data:
        row_arry = []
        for val in row:
            row_arry.append(val.value)
        data_arr.append(row_arry)

    #Creating the approach egress pairs for dataframe header
    approach_egress = []
    for idx, val in enumerate(egress):
        approach_idx = math.floor(idx/3)
        approach_egress.append(f"{approach[approach_idx]}-{val}")
    approach_egress

    #Creating the dataframe
    df = pd.DataFrame(data_arr, columns=approach_egress, index=vehicles)

    #Getting the index and colums for navigating the dataframe using loc
    a_e_headers = df.columns.values.tolist()
    vehicles = df.index.values.reshape(1,-1)
    date = pd.to_datetime(date.value).date()

    #Connects to the server
    conn = sqlite3.connect(database_path)
    cur = conn.cursor()

    for vehicle in vehicles[0]:    #Adds the data
        vehicle = vehicle[0]
        # vehicle = vehicle[0][0]

        for a_e in a_e_headers:
            count = float(df.loc[vehicle,a_e])
            app = a_e.split('-')[0]
            egg = a_e.split('-')[1]
            cur.execute(f"INSERT INTO data VALUES ('{project_id}', '{date}', {start_hour.value}, {end_hour.value}, '{app}', '{egg}', 'intersection','{vehicle}', {count})")

    conn.commit()
    conn.close()

def threeleg_intersection_to_db(wb_path, database_path):
    """
    This function reads data from 4-Leg intersection Excel file and inserts it into a SQLite database.

    Parameters:
        wb_path (str): The path to the Excel file.
        database_path (str): The path to the SQLite database.

    Returns:
        None
    """
    wb = xl.load_workbook(wb_path)
    ws = wb.active


    #Get the start Hour
    start_hour  = ws['C2']
    end_hour    = ws['E2']
    project_id  = ws['B1'].value
    date        = ws['G2']
    approach    = [i.value for i in ws['B3:G3'][0] if i.value != None]
    egress      = [i.value for i in ws['B4:G4'][0]]
    vehicles    = [[i[0].value for i in ws['A5:A12']]]
    data        = ws['B5:G12']

    #Transforms the openpyxl cell to actual data
    data_arr = []
    for row in data:
        row_arry = []
        for val in row:
            row_arry.append(val.value)
        data_arr.append(row_arry)

    #Creating the approach egress pairs for dataframe header
    approach_egress = []
    for idx, val in enumerate(egress):
        approach_idx = math.floor(idx/2)
        approach_egress.append(f"{approach[approach_idx]}-{val}")
    approach_egress

    #Creating the dataframe
    df = pd.DataFrame(data_arr, columns=approach_egress, index=vehicles)

    #Getting the index and colums for navigating the dataframe using loc
    a_e_headers = df.columns.values.tolist()
    vehicles = df.index.values.reshape(1,-1)
    date = pd.to_datetime(date.value).date()

    #Connects to the server
    conn = sqlite3.connect(database_path)
    cur = conn.cursor()

    for vehicle in vehicles[0]:    #Adds the data
        vehicle = vehicle[0]
        # vehicle = vehicle[0][0]

        for a_e in a_e_headers:
            count = float(df.loc[vehicle,a_e])
            app = a_e.split('-')[0]
            egg = a_e.split('-')[1]
            cur.execute(f"INSERT INTO data VALUES ('{project_id}', '{date}', {start_hour.value}, {end_hour.value}, '{app}', '{egg}', 'intersection','{vehicle}', {count})")

    conn.commit()
    conn.close()

def midblock_to_db(wb_path, database_path):
    """
    Inserts data from a Midblock Excel file into a SQLite database.

    Parameters:
        wb_path (str): The path to the Excel file.
        database_path (str): The path to the SQLite database.

    Returns:
        None
    """
    wb = xl.load_workbook(wb_path)
    ws = wb.active

    #Get the cell data
    project_id = ws['C1'].value
    date = pd.to_datetime(ws['E1'].value)
    direction = ws['G1'].value
    vehicles = [i.value for i in ws['C2:J2'][0]]
    data = ws['C3:J26']

    #Transforms the openpyxl cell to actual data
    data_arr = []
    for row in data:
        row_arry = []
        for val in row:
            row_arry.append(val.value)
        data_arr.append(row_arry)

    #Generates time pairs with example "1-2" for the whole day as index
    time_pairs = [f"{start}-{end}" for start, end in zip(np.arange(0,24), np.hstack([np.arange(1,24),0]))]
    df = pd.DataFrame(data_arr, index=time_pairs, columns=vehicles)

    #Connects to the server
    conn = sqlite3.connect(database_path)
    cur = conn.cursor()

    for pair in time_pairs: #Commits the data
        for vehicle in vehicles:
            count = float(df.loc[pair, vehicle])
            app, egg = direction.split('-')
            start_hour, end_hour = pair.split('-')
            cur.execute(f"INSERT INTO data VALUES ('{project_id}', '{date}', {start_hour}, {end_hour}, '{app}', '{egg}', 'midblock','{vehicle}', {count})")
        
    conn.commit()
    conn.close()

def pcef_to_db(wb_path, database_path):
    """
    Inserts data from a PCEF Excel file into a SQLite database.

    Parameters:
        wb_path (str): The path to the Excel file.
        database_path (str): The path to the SQLite database.

    Returns:
        None
    """
    df = pd.read_excel(wb_path)

    #Connects to the server
    conn = sqlite3.connect(database_path)
    cur = conn.cursor()

    data = df.to_numpy()
    for row in data:
        cur.execute(f"INSERT INTO vehicle VALUES ('{row[0]}', {row[1]})")
    conn.commit()
    conn.close()

    return True

def capacity_to_db(wb_path, database_path):
    """
    Inserts data from a Capacity Excel file into a SQLite database.

    Parameters:
        wb_path (str): The path to the Excel file.
        database_path (str): The path to the SQLite database.

    Returns:
        None
    """
    df = pd.read_excel(wb_path)

    #Connects to the server
    conn = sqlite3.connect(database_path)
    cur = conn.cursor()

    data = df.to_numpy()
    for row in data:
        cur.execute(f"INSERT INTO lane_properties VALUES ('{row[0]}', {row[1]}, {row[2]}, '{row[3]}')")
    
    conn.commit()
    conn.close()

    return True